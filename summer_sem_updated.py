# on the cmd, run the file as follows:
# python filename.py <xlsx_file_name_with_extension> <name_of_sheet_inside_the_xlsx_file_for_input> <day (enter as Monday, Tuesday, ie full and capitalised name of the day)>

import os
import openpyxl
import sys
import collections
from openpyxl.styles import Font
# baking required variables
day_slots = {
    'Monday': ['A11+A12', 'B11+B12', 'C21+C22', 'D21+D22'],
    'Tuesday': ['B11+B12', 'A11+A12', 'D21+D22', 'C21+C22'],
    'Wednesday': ['A11+A12', 'B11+B12', 'C21+C22', 'D21+D22'],
    'Thursday': ['B11+B12', 'A11+A12', 'D21+D22', 'C21+C22'],
    'Friday': ['TB11+TB12', 'TA11+TA12', 'TD21+TD22', 'TC21+TC22']
}

output_file = openpyxl.Workbook()


# the CLI arguments
input_xlsx_file = sys.argv[1]
input_xlsx_sheet = sys.argv[2]
day = sys.argv[3]

# print(day)
input_wb = openpyxl.load_workbook(input_xlsx_file)
input_sheet = input_wb[input_xlsx_sheet] # selecting the sheet from loaded workbook


rows = input_sheet.max_row
cols = input_sheet.max_column

slots_for_the_day = day_slots[day]

output_sheet = output_file.create_sheet(day) # creating sheet to be created in the output_xlsx_file
output_sheet['A1'] = 'Venue'

for i in range(1,cols+1):
    output_sheet.cell(row = 1, column = i).font = Font(bold=True)

i=2
for slot in day_slots[day]: # the order of slots appearing in the output_xlsx_sheet must be in the same order as in Monday
    output_sheet.cell(row = 1, column = i).value = slot
    i += 1

out_column = 1

out_dict = {}

for slot in slots_for_the_day:
    # out_row = 2
    # out_column += 1
    for i in range(2,rows+1):
        # out_column = 1
        if slot == input_sheet.cell(row = i, column = 12).value:
            venue = input_sheet.cell(row = i, column = 11).value

            if venue in out_dict:
                out_dict[venue].append({slot: input_sheet.cell(row = i, column = 2).value + ' - ' + input_sheet.cell(row = i, column = 14).value})
            else:
                out_dict[venue] = []
                out_dict[venue].append({slot: input_sheet.cell(row = i, column = 2).value + ' - ' + input_sheet.cell(row = i, column = 14).value})

# print(out_dict)



i=2
# checking_slot = 2
for key in out_dict:
    # print(key)
    checking_slot = 2
    output_sheet.cell(row = i, column = 1).value = key
    list_of_slot_course_dict = out_dict[key]
    for slot in day_slots['Monday']:
        flag = 0
        for item in list_of_slot_course_dict:
            if slot in item.keys():
                output_sheet.cell(row = i, column = checking_slot).value = item[slot]
                checking_slot += 1
                flag = 1
                break

        if flag == 0:
            output_sheet.cell(row = i, column = checking_slot).value = '-'
    i += 1

output_sheet.column_dimensions['B'].width = (len(output_sheet['B2'].value) + 2) * 1.2
output_sheet.column_dimensions['C'].width = (len(output_sheet['C2'].value) + 2) * 1.2
output_sheet.column_dimensions['D'].width = (len(output_sheet['D2'].value) + 2) * 1.2
output_sheet.column_dimensions['E'].width = (len(output_sheet['E2'].value) + 2) * 1.2

save_name = day +'.xlsx'

if os.path.isfile(save_name):
    os.remove(save_name)

output_file.save(save_name)
