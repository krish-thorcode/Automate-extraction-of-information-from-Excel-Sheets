# on the cmd, run the file as follows:
# python filename.py <xlsx_file_name_with_extension> <name_of_sheet_inside_the_xlsx_file_for_input> <day (enter as Monday, Tuesday, ie full and
#capitalised name of the day)>

import openpyxl
import sys
import collections
import os
from openpyxl.styles import Font
# baking required variables
day_slots = {
    'Monday': ['A1','F1','D1','TB1','TG1','A2','F2','D2','TB2','TG2'],
    'Tuesday': ['B1','G1','E1','TC1','TAA1','B2','G2','E2','TC2','TAA2'],
    'Wednesday': ['C1','A1','F1','V1','V2','C2','A2','F2','TD2','TBB2'],
    'Thursday': ['D1','B1','G1','TE1','TCC1','D2','B2','G2','TE2','TCC2'],
    'Friday': ['E1','C1','TA1','TF1','TD1','E2','C2','TA2','TF2','TD2']
    # 'Saturday': ['V8','X1','Y1','X2','Z','Y2','W2','V9'],
    # 'Sunday': ['V10','Y1','X1','Y2','Z','X2','W2','V11']
}

# day_timing_slots = {
#     'Monday': [' 8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50'],
#     'Tuesday': [' 8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50'],
#     'Wednesday': [' 8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50'],
#     'Thursday': [' 8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50'],
#     'Friday': [' 8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50']
#     # 'Saturday': [' 8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50'],
#     # 'Sunday': [' 8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50'],
# }
day_timing_slots = ['8.00 - 8.50','9.00 - 9.50','10.00 - 10.50','11.00  11.50','12.00 - 12.50','2.00 - 2.50','03.00 - 3.50','4.00 - 4.50','5.00 - 5.50','6.00 - 6.50']

output_file = openpyxl.Workbook()

# print(sys.argv)
# the CLI arguments
input_xlsx_file = sys.argv[1]
input_xlsx_sheet = sys.argv[2]
day = sys.argv[3]

print(input_xlsx_file)
print(input_xlsx_sheet)
print(day)

# print(day)
input_wb = openpyxl.load_workbook(input_xlsx_file)
input_sheet = input_wb[input_xlsx_sheet] # selecting the sheet from loaded workbook
# input_sheet = input_wb[input_xlsx_sheet]

rows = input_sheet.max_row
cols = input_sheet.max_column

slots_for_the_day = day_slots[day]

output_sheet = output_file.create_sheet(day) # creating sheet to be created in the output_xlsx_file
output_sheet['A1'] = 'Venue'

for i in range(1,cols+1):
    output_sheet.cell(row = 1, column = i).font = Font(bold=True)

i=2
iterate = 0
for slot in day_slots[day]: # the order of slots appearing in the output_xlsx_sheet must be in the same order as in the day entered as arg
    output_sheet.cell(row = 1, column = i).value = slot + ' (' + day_timing_slots[iterate] + ')'
    iterate += 1
    i += 1

out_column = 1

out_dict = {}

for slot in slots_for_the_day:
    for i in range(2,rows+1):
        if slot in input_sheet.cell(row = i, column = 12).value.split('+'):
            # print('iron man')
            venue = input_sheet.cell(row = i, column = 11).value

            if venue in out_dict:
                out_dict[venue].append({slot: input_sheet.cell(row = i, column = 2).value + ' - ' + input_sheet.cell(row = i, column = 14).value})
                # if slot == 'TC2':
                #     print(venue, input_sheet.cell(row = i, column = 2).value, input_sheet.cell(row = i, column = 14).value, slot)
            else:
                out_dict[venue] = []
                out_dict[venue].append({slot: input_sheet.cell(row = i, column = 2).value + ' - ' + input_sheet.cell(row = i, column = 14).value})

# print(*out_dict['SJTG01'],sep = '\n ')


out_dict = collections.OrderedDict(sorted(out_dict.items()))
# print(out_dict)


i=2
# checking_slot = 2
for key in out_dict: # each time key is assigned to a venue
    # print(key)
    checking_slot = 2
    output_sheet.cell(row = i, column = 1).value = key
    list_of_dict_slot_and_course_code_faculty = out_dict[key] # for the present 'key' (venue), list of course code and corresponding faculty dictionary which are identified by slots
    # print(list_of_dict_slot_and_course_code_faculty)
    for slot in day_slots[day]:
        flag = 0
        for item in list_of_dict_slot_and_course_code_faculty: # item will be a key: value pair, key = slot, value = course code-faculty
            # print(item)
            if slot in item.keys(): # since each list item is a dictionary with a single key-pair value, item.keys() return only one value
                output_sheet.cell(row = i, column = checking_slot).value = item[slot]
                checking_slot += 1
                flag = 1
                break

        if flag == 0:
            output_sheet.cell(row = i, column = checking_slot).value = '-'
            checking_slot += 1
    i += 1

def set_column_width():
    cols_output = output_sheet.max_column
    rows_output = output_sheet.max_row
    maxlen_in_col = 0
    ch = 'A'
    for i in range(0,cols_output):
        # print(ch)
        col = output_sheet[ch] # the entire column is copied to the col variable as a list of cell objects
        if col[0].value:
            maxlen_in_col = len(col[0].value)
        for j in range(1,rows_output):
            if col[j].value:
                if maxlen_in_col < len(col[j].value):
                    maxlen_in_col = len(col[j].value)

        output_sheet.column_dimensions[ch].width = (maxlen_in_col)*1.2
        ch = chr(ord(ch)+1)

set_column_width()

save_name = day +'.xlsx'

if os.path.isfile(save_name):
    os.remove(save_name)

output_file.save(save_name)


            # output_sheet.cell(row = out_row, column = out_column++).value = input_sheet.cell(row = i, column = 11).value
            # output_sheet.cell(row = out_row, column = out_column).value = input_sheet.cell(row = i, column = 2).value + ' - '
            # output_sheet.cell(row = out_row++, column = out_column).value += input_sheet.cell(row = i, column = 14).value
